import datetime
import logging
from copy import deepcopy

import openpyxl
from cds.models import (
    DidatticaCds,
    DidatticaCdsCollegamento,
    DidatticaCdsLingua,
    DidatticaClasseLaurea,
    DidatticaRegolamento,
)
from cds_brochure.models import (
    CdsBrochure,
    CdsBrochureExStudenti,
    CdsBrochureLink,
    CdsBrochureSlider,
)
from django.contrib import messages
from django.contrib.admin.models import ADDITION, CHANGE
from django.db import DatabaseError, IntegrityError, OperationalError, transaction
from django.db.models import Q
from django.utils.translation import gettext_lazy as _
from generics.utils import get_latest_available_dummy_id, log_action_on_commit
from import_assistant.exceptions import (
    InvalidValueError,
    BaseImportError,
    EntryImportError,
    MissingHeadersError,
    MissingValueError,
)
from import_assistant.settings import (
    ACCEPTED_TIPO_CORSO_COD,
    CDS_IMPORT_FIELDS,
    EXCEL_FIELDS_MAPPINGS,
)
from structures.models import DidatticaDipartimento

logger = logging.getLogger(__name__)


# ------------------------
# Utility and Helper Functions
# ------------------------


def _handle_entry_import_error(request, error, entry_id, entry_name):
    """Helper function to handle entry import errors."""
    error.set_entry_name(entry_name)
    error.set_entry_id(entry_id)
    error.skip_entry()
    logger.error(str(error))
    messages.error(request, str(error))


def _create_or_update_brochure(request, data, academic_year, cds):
    """
    Updates an existing brochure or creates a new one if none exists.

    If a brochure already exists but corresponds to a different academic year,
    it updates the year and logs the action. Otherwise, it creates a new
    brochure entry, potentially using data from a previous brochure.
    """
    numero_posti = data.get(EXCEL_FIELDS_MAPPINGS["NUMERO_POSTI"], None)
    cds_brochure = CdsBrochure.objects.filter(cds_cod=cds.cds_cod).first()
    if cds_brochure is not None:
        # Update the existing brochure's academic year and log the change
        if cds_brochure.aa != academic_year or cds_brochure.num_posti != numero_posti:
            cds_brochure.aa = academic_year
            cds_brochure.user_mod = request.user
            cds_brochure.dt_mod = datetime.datetime.now()

            if numero_posti:
                cds_brochure.num_posti = numero_posti

            cds_brochure.save()

            log_action_on_commit(
                request.user,
                cds_brochure,
                CHANGE,
                _("Updated brochure through import procedure for cds [{}]").format(
                    cds.cds_cod
                ),
            )

        return

    # Create a new brochure using previous data if available
    previous_cds_brochure = _get_previous_cds_brochure(data)
    cds_brochure = (
        deepcopy(previous_cds_brochure) if previous_cds_brochure else CdsBrochure()
    )

    cds_brochure.id = None
    cds_brochure.cds = cds
    cds_brochure.cds_cod = cds.cds_cod
    cds_brochure.aa = academic_year
    cds_brochure.user_mod = request.user
    cds_brochure.dt_mod = datetime.datetime.now()

    if numero_posti:
        cds_brochure.num_posti = numero_posti

    cds_brochure.save()

    log_action_on_commit(
        request.user,
        cds_brochure,
        ADDITION,
        _("Created brochure for cds [{}]").format(cds.cds_cod),
    )

    # Import related data (sliders, links, ex-students) from the previous brochure
    if previous_cds_brochure:
        _import_related_data(request, cds_brochure, previous_cds_brochure)


def _import_related_data(request, cds_brochure, previous_cds_brochure):
    """Imports related slider, ex-student, and link data from previous brochures"""
    _import_previous_entries(
        request, cds_brochure, previous_cds_brochure, CdsBrochureSlider, "slider"
    )
    _import_previous_entries(
        request, cds_brochure, previous_cds_brochure, CdsBrochureExStudenti, "exstudent"
    )
    _import_previous_entries(
        request, cds_brochure, previous_cds_brochure, CdsBrochureLink, "link"
    )


def _import_previous_entries(
    request, cds_brochure, previous_cds_brochure, model, entry_name
):
    """
    Imports specific related data entries from a previous brochure.
    """
    previous_entries = model.objects.filter(cds_brochure=previous_cds_brochure)
    for entry in previous_entries:
        # Clone the previous entry, update references, and save the new instance
        new_entry = deepcopy(entry)
        new_entry.id = None
        new_entry.cds_brochure = cds_brochure
        new_entry.save()
        log_action_on_commit(
            request.user,
            cds_brochure,
            CHANGE,
            _("Imported {} from previous years").format(entry_name),
        )


def _get_previous_cds_brochure(data):
    """
    Retrieves the most recent previous brochure based on the CDS code provided in the data.
    """
    previous_cds_cod = data.get(EXCEL_FIELDS_MAPPINGS["CDS_COD_PREC"])
    if previous_cds_cod:
        previous_cds = DidatticaCds.objects.filter(cds_cod=previous_cds_cod).first()
        if previous_cds:
            return CdsBrochure.objects.filter(cds=previous_cds).order_by("-aa").first()
    return None


def _create_regdid_if_missing(request, data, academic_year, cds, regdid_id):
    """
    Creates a new DidatticaRegolamento (RegDid) entry if one does not already exist for the academic year.

    If a RegDid for the same CDS and academic year exists, it is skipped.
    Otherwise, it creates a new RegDid entry based on a previous one (if available)
    or initializes a new entry with default values.
    """
    # Check for previous CDS linkage
    cds_collegamento_prec = DidatticaCdsCollegamento.objects.filter(cds=cds).first()
    regdid_cds_lookup_query = Q(cds=cds)
    if cds_collegamento_prec:
        regdid_cds_lookup_query |= Q(cds=cds_collegamento_prec.cds_prec)

    # Look for an existing RegDid for the CDS or linked CDS
    regdid_prec = (
        DidatticaRegolamento.objects.filter(regdid_cds_lookup_query)
        .order_by("-aa_reg_did")
        .first()
    )

    if regdid_prec:
        # Skip if a RegDid for the same academic year already exists
        if regdid_prec.cds == cds and regdid_prec.aa_reg_did == academic_year:
            return

        # Create a new RegDid based on the previous one
        regdid = deepcopy(regdid_prec)
        regdid.regdid_id = regdid_id
        regdid.cds = cds
        regdid.aa_reg_did = academic_year
        regdid.stato_regdid_cod = "A"
        regdid.stato_regdid_des = "Attivo"

        # Set modalita erogazione if provided
        if data.get(EXCEL_FIELDS_MAPPINGS["MODALITA_EROGAZIONE"]):
            regdid.modalita_erogazione = data.get(
                EXCEL_FIELDS_MAPPINGS["MODALITA_EROGAZIONE"]
            )

        regdid.save()

    else:
        # Create a completely new RegDid
        regdid = DidatticaRegolamento.objects.create(
            regdid_id=regdid_id,
            aa_reg_did=academic_year,
            cds=cds,
            modalita_erogazione=data.get(EXCEL_FIELDS_MAPPINGS["MODALITA_EROGAZIONE"]),
            stato_regdid_cod="A",
            stato_regdid_des="Attivo",
        )

    log_action_on_commit(
        request.user, regdid, CHANGE, _("Created RegDid for cds {}").format(cds.cds_cod)
    )


def _create_cds_collegamento_if_missing(request, cds, cds_cod_prec):
    """
    Creates a new DidatticaCdsCollegamento entry if one does not already exist.

    Links the current CDS to a previous CDS if a valid previous code is provided.
    """
    # Skip if no previous code is provided or a link already exists
    if (
        not cds_cod_prec
        or DidatticaCdsCollegamento.objects.filter(
            Q(cds_prec__cds_cod=cds_cod_prec) | Q(cds__cds_cod=cds.cds_cod)
        ).exists()
    ):
        return

    # Retrieve the previous CDS by its code. Since cds_cod is not unique, it orders them by the latest related regdid,
    # if not available, it orders them by the latest cds_id and picks the first one
    cds_prec = (
        DidatticaCds.objects.filter(cds_cod=cds_cod_prec)
        .order_by("-didatticaregolamento__aa_reg_did", "-cds_id")
        .first()
    )
    if not cds_prec:
        raise InvalidValueError(
            field_name={EXCEL_FIELDS_MAPPINGS["CDS_COD_PREC"]}, value=cds_cod_prec
        )

    # Create the linkage between current CDS and previous CDS
    CDS_COLLEGAMENTO = DidatticaCdsCollegamento.objects.create(
        cds=cds, cds_prec=cds_prec
    )
    log_action_on_commit(
        request.user,
        CDS_COLLEGAMENTO,
        ADDITION,
        _("Linked {} to {}").format(
            CDS_COLLEGAMENTO.cds.cds_cod, CDS_COLLEGAMENTO.cds_prec.cds_cod
        ),
    )


def _create_cds_languages_if_missing(request, data, cds):
    """
    Creates language entries for a CDS if they do not already exist.

    Validates the provided language codes and creates new DidatticaCdsLingua entries
    for each language in the provided data.
    """
    langs = data.get(EXCEL_FIELDS_MAPPINGS["LINGUE"])
    langs = langs.lower()
    langs = langs.replace(" ", "")
    langs = langs.split(",")

    # Validate language codes
    if len(langs) < 1 or any(lang not in ("ita", "eng") for lang in langs):
        raise InvalidValueError(
            field_name={data.get(EXCEL_FIELDS_MAPPINGS["LINGUE"])},
            value=data.get(EXCEL_FIELDS_MAPPINGS["LINGUE"]),
        )

    # Create language entries if missing
    for lang in langs:
        if DidatticaCdsLingua.objects.filter(
            cdsord_id=cds.cdsord_id, iso6392_cod=lang
        ).exists():
            continue
        lin_did_ord_id = get_latest_available_dummy_id(
            DidatticaCdsLingua, "lin_did_ord_id"
        )
        cds_lingua = DidatticaCdsLingua.objects.create(
            lin_did_ord_id=lin_did_ord_id,
            cdsord=cds,
            lingua_id=1 if lang == "ita" else 2,
            lingua_des_it="ITALIANO" if lang == "ita" else "INGLESE",
            iso6392_cod=lang,
            lingua_des_eng="Italian" if lang == "ita" else "English",
        )
        log_action_on_commit(
            request.user,
            cds_lingua,
            ADDITION,
            _("Created language: {} for cds {}").format(lang, cds.cds_cod),
        )


def _validate_tipo_corso_cod(data):
    if (
        data.get(EXCEL_FIELDS_MAPPINGS["TIPO_CORSO_COD"]).upper()
        not in ACCEPTED_TIPO_CORSO_COD
    ):
        raise InvalidValueError(
            field_name=EXCEL_FIELDS_MAPPINGS["TIPO_CORSO_COD"],
            value=data.get(EXCEL_FIELDS_MAPPINGS["TIPO_CORSO_COD"]),
        )


def _get_or_create_cds(request, data, academic_year, cds_cod, cds_id):
    """
    Retrieves an existing CDS or creates a new one if not found.
    """
    _validate_tipo_corso_cod(data)

    cds = DidatticaCds.objects.filter(cds_cod=cds_cod).first()
    if not cds:
        # Retrieve the department associated with the CDS
        department = DidatticaDipartimento.objects.filter(
            dip_cod=data.get(EXCEL_FIELDS_MAPPINGS["DIP_COD"])
        ).first()

        # Retrieve the DidatticaClasseLaurea associated with the CDS for cla_miur_cod and intercla_miur_cod
        cla_m = DidatticaClasseLaurea.objects.filter(
            cla_miur_cod=data.get(EXCEL_FIELDS_MAPPINGS["CLA_MIUR_COD"]).upper()
        ).first()
        if cla_m is None:
            raise InvalidValueError(
                field_name=EXCEL_FIELDS_MAPPINGS["CLA_MIUR_COD"],
                value=data.get(EXCEL_FIELDS_MAPPINGS["CLA_MIUR_COD"]),
            )
        cla_m_id = cla_m.cla_m_id

        intercla_m_id = None
        if data.get(EXCEL_FIELDS_MAPPINGS["INTERCLA_MIUR_COD"]):
            intercla_m = DidatticaClasseLaurea.objects.filter(
                cla_miur_cod=data.get(
                    EXCEL_FIELDS_MAPPINGS["INTERCLA_MIUR_COD"]
                ).upper()
            ).first()
            if intercla_m is None:
                raise InvalidValueError(
                    field_name=EXCEL_FIELDS_MAPPINGS["INTERCLA_MIUR_COD"],
                    value=data.get(EXCEL_FIELDS_MAPPINGS["INTERCLA_MIUR_COD"]),
                )
            intercla_m_id = intercla_m.intercla_m_id

        # Create a new CDS entry
        cds = DidatticaCds.objects.create(
            cds_id=cds_id,
            cds_cod=cds_cod,
            nome_cds_it=data.get(EXCEL_FIELDS_MAPPINGS["NOME_CDS_IT"]),
            nome_cds_eng=data.get(EXCEL_FIELDS_MAPPINGS["NOME_CDS_ENG"]),
            tipo_corso_cod=data.get(EXCEL_FIELDS_MAPPINGS["TIPO_CORSO_COD"]).upper(),
            tipo_corso_des=data.get(EXCEL_FIELDS_MAPPINGS["TIPO_CORSO_DES"]),
            cla_m_id=cla_m_id,
            cla_miur_cod=data.get(EXCEL_FIELDS_MAPPINGS["CLA_MIUR_COD"]).upper(),
            cla_miur_des=data.get(EXCEL_FIELDS_MAPPINGS["CLA_MIUR_DES"]),
            intercla_m_id=intercla_m_id,
            intercla_miur_cod=data.get(
                EXCEL_FIELDS_MAPPINGS["INTERCLA_MIUR_COD"]
            ).upper(),
            intercla_miur_des=data.get(EXCEL_FIELDS_MAPPINGS["INTERCLA_MIUR_DES"]),
            durata_anni=data.get(EXCEL_FIELDS_MAPPINGS["DURATA_ANNI"]),
            valore_min=data.get(EXCEL_FIELDS_MAPPINGS["VALORE_MIN"]),
            dip=department,
            aa_ord_id=academic_year,
            stato_cdsord_cod="A",
            cdsord_id=cds_id,
        )
        log_action_on_commit(
            request.user, cds, CHANGE, _("Imported CdS [{}]").format(cds_cod)
        )

    return cds


def _validate_mandatory_fields(data, required_fields):
    """Validates that all required values are present in the data"""
    for field in required_fields:
        if not data.get(field):
            raise MissingValueError(field_name=field)


def _validate_headers(required_headers, headers):
    """Checks for required fields in headers and displays an error message if missing"""
    missing_headers = [field for field in required_headers if field not in headers]
    if missing_headers:
        raise MissingHeadersError(headers=missing_headers)


# ------------------------
# Main Import Logic
# ------------------------


def handle_cds_import(request, file, academic_year):
    """
    Handles the import of study courses (CDS) from an Excel file.

    This function performs the following:
    - Reads the Excel file to extract data rows.
    - Validates headers and mandatory fields.
    - Processes each row to create or update CDS entries and related data (eg. DidatticaCdsLingua, DidatticaCdsCollegamento, DidatticaRegolamento, CdsBrochure).
    """
    # Results
    total_rows = 0
    skipped_rows = 0

    wb = openpyxl.load_workbook(filename=file, read_only=True, data_only=True)
    ws = wb.active

    if ws.max_row < 2:
        raise BaseImportError(_("At least one entry is required"))

    headers = [cell.value for cell in ws[1]]

    REQUIRED_FIELDS = [
        field_properties["name"]
        for field_properties in CDS_IMPORT_FIELDS
        if field_properties["required"]
    ]

    REQUIRED_HEADERS = list(EXCEL_FIELDS_MAPPINGS.values())

    _validate_headers(REQUIRED_HEADERS, headers)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            total_rows += 1
            if all(cell is None for cell in row):
                raise EntryImportError("Empty")

            data = dict(zip(headers, row))
            try:
                _validate_mandatory_fields(data, REQUIRED_FIELDS)
            except EntryImportError as e:
                skipped_rows += 1
                _handle_entry_import_error(request, e, row_idx, "Row")
                continue

            latest_cds_id = get_latest_available_dummy_id(DidatticaCds, "cds_id")
            latest_reg_id = get_latest_available_dummy_id(
                DidatticaRegolamento, "regdid_id"
            )
            cds_cod = data.get(EXCEL_FIELDS_MAPPINGS["CDS_COD"]) or latest_cds_id
            cds_cod_prec = data.get(EXCEL_FIELDS_MAPPINGS["CDS_COD_PREC"])

            try:
                with transaction.atomic():
                    cds = _get_or_create_cds(
                        request, data, academic_year, cds_cod, latest_cds_id
                    )
                    _create_cds_collegamento_if_missing(request, cds, cds_cod_prec)
                    _create_cds_languages_if_missing(request, data, cds)
                    _create_regdid_if_missing(
                        request, data, academic_year, cds, latest_reg_id
                    )
                    _create_or_update_brochure(request, data, academic_year, cds)
            except (IntegrityError, OperationalError, DatabaseError) as e:
                skipped_rows += 1
                logger.error(_("Database error during transaction: {}").format(e))
        except EntryImportError as e:
            skipped_rows += 1
            _handle_entry_import_error(request, e, row_idx, "Row")

    return (total_rows, skipped_rows)
