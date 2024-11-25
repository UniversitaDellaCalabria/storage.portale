from io import BytesIO

import openpyxl
from cds.models import DidatticaCds
from django.conf import settings
from django.db.models import F
from import_assistant.settings import (
    ACCEPTED_TIPO_CORSO_COD,
    EXCEL_FIELDS_MAPPINGS,
)


def gen_xlsx_for_cds_export():
    """
    Returns an excel representation of the expected file used for CDS import,
    populated with data for settings.CURRENT_YEAR
    """
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = str(settings.CURRENT_YEAR)
    HEADERS = list(EXCEL_FIELDS_MAPPINGS.values())
    for col_num, header in enumerate(HEADERS, start=1):
        sheet.cell(row=1, column=col_num, value=header)

    data = (
        DidatticaCds.objects.filter(
            didatticaregolamento__aa_reg_did=settings.CURRENT_YEAR,
            didatticaregolamento__stato_regdid_cod="A",
            tipo_corso_cod__in=ACCEPTED_TIPO_CORSO_COD,
        )
        .prefetch_related("didatticacdslingua")
        .annotate(
            **{
                EXCEL_FIELDS_MAPPINGS["CDS_COD_PREC"]: F(
                    "cds_collegamento__cds_prec__cds_cod"
                ),
                EXCEL_FIELDS_MAPPINGS["CDS_COD"]: F("cds_cod"),
                EXCEL_FIELDS_MAPPINGS["NOME_CDS_IT"]: F("nome_cds_it"),
                EXCEL_FIELDS_MAPPINGS["NOME_CDS_ENG"]: F("nome_cds_eng"),
                EXCEL_FIELDS_MAPPINGS["DIP_COD"]: F("dip__dip_cod"),
                EXCEL_FIELDS_MAPPINGS["TIPO_CORSO_COD"]: F("tipo_corso_cod"),
                EXCEL_FIELDS_MAPPINGS["TIPO_CORSO_DES"]: F("tipo_corso_des"),
                EXCEL_FIELDS_MAPPINGS["DURATA_ANNI"]: F("durata_anni"),
                EXCEL_FIELDS_MAPPINGS["VALORE_MIN"]: F("valore_min"),
                EXCEL_FIELDS_MAPPINGS["CLA_MIUR_COD"]: F("cla_miur_cod"),
                EXCEL_FIELDS_MAPPINGS["CLA_MIUR_DES"]: F("cla_miur_des"),
                EXCEL_FIELDS_MAPPINGS["INTERCLA_MIUR_COD"]: F("intercla_miur_cod"),
                EXCEL_FIELDS_MAPPINGS["INTERCLA_MIUR_DES"]: F("intercla_miur_des"),
                EXCEL_FIELDS_MAPPINGS["MODALITA_EROGAZIONE"]: F(
                    "didatticaregolamento__modalita_erogazione"
                ),
                EXCEL_FIELDS_MAPPINGS["NUMERO_POSTI"]: F("brochure__num_posti"),
            }
        )
        .order_by("cds_cod")
    )

    # Write data to the Excel sheet
    row_num = 2  # Start from the second row for data
    for record in data:
        for col_num, field in enumerate(HEADERS, start=1):
            # Dynamically fetch the value for each header from the record
            field_name = next(
                value for _, value in EXCEL_FIELDS_MAPPINGS.items() if value == field
            )
            sheet.cell(
                row=row_num, column=col_num, value=getattr(record, field_name, "")
            )

        # Handle languages separately
        field_name = EXCEL_FIELDS_MAPPINGS["LINGUE"]
        col_num = HEADERS.index(field_name) + 1
        lingue = ",".join(
            record.didatticacdslingua.order_by("lingua_id").values_list(
                "iso6392_cod", flat=True
            )
        )
        sheet.cell(row=row_num, column=col_num, value=lingue)

        row_num += 1

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    return output
